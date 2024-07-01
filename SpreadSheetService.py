from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import gspread
import datetime
import heapq
from openpyxl.utils import get_column_letter


api_path = r"C:\Users\sabax\Documents\backend\apitest-418907-1658bd7509d0.json"
folder_id = '163LOPPPAgKUZXpegrYbBbZkGntFnH5-v'

#google driveに設定
scopes = ['https://www.googleapis.com/auth/drive']

credentials = Credentials.from_service_account_file(
    api_path,
    scopes = scopes
)

class SpreadsheetService:
    def __init__(self, fileID=None):
        credentials = Credentials.from_service_account_file(
            api_path,
            scopes = scopes
        )
        self.credentials = credentials
        self.fileID = fileID
        self.gc = gspread.authorize(credentials)
        #Google Drive APIに接続
        self.drive_service = build('drive', 'v3', credentials=credentials)
        # Google Sheets APIに接続
        self.sheets_service = build('sheets', 'v4', credentials=credentials)
        self.query = f"'{folder_id}' in parents"

    def get_info(self):
        """
        スプレッドシートのファイル名とIDを取得
        """
        try:
            results = self.drive_service.files().list(q=self.query, fields='files(name, id)').execute()
            items = results.get('files', [])
            return items
        except Exception as e:
            print(e)


    def get_worksheet(self):
        """
        sheetidのシート名とIDを辞書として返す
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        worksheets = spreadsheet.worksheets()
        return [{'label': ws.title, 'value': ws.id} for ws in worksheets]


    def get_date_info(self, sheetname):
        """
        日付の入った行の値を取得
        returnは配列
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sheetname)
        compar_day = spreadsheet.row_values(1)
        print(compar_day)


    def get_closest_positions(self, dates_positions, target_date):
        """
        今日から最も日付が近いセルの位置を取得
        dict型 'row': 列番号, 'str_row' : '文字型の列番号(ex ABG etc)'
        """
        format_str = "%m/%d %H:%M"
        differences = [abs(datetime.datetime.strptime(target_date, format_str) - datetime.datetime.strptime(date_position[0], format_str)) for date_position in dates_positions]
        closest_indices = heapq.nsmallest(3, range(len(differences)), key=differences.__getitem__)
        return [(dates_positions[i][1]) for i in closest_indices]


    def closetDataFinder(self, sheetname):
        """
        取得した日付と近い順に３つ値と位置を取得する
        returnはint 位置のみ
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sheetname)

        format_str = "%m/%d %H:%M"
        dates_positions = []
        row = 2

        while True:
            compar_day = spreadsheet.cell(1, row).value
            if compar_day is None:
                break

            f_format_day = datetime.datetime.strptime(compar_day, format_str).strftime(format_str)
            dates_positions.append({'value':f_format_day, 'position':row})
            row += 42  # 4枚スキップ

        # 選択したデータの近辺をさらに取得
        return self.near_compar_date(sheetname=sheetname, row_data=dates_positions)


    def near_compar_date(self, sheetname, row_data):
        """
        日付に関してのデータの取得
        closetDataFinderで得たデータ付近を探索して日付が近いデータの位置を返す
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sheetname)
        dates_positions = []

        format_str = "%m/%d %H:%M"
        try:
            max_position = max(item['position'] for item in row_data)    # dates_positionsの最大値を取得する

            for i in range(0, 35, 7):
                now_position = max_position + i
                compar_day = spreadsheet.cell(1, now_position).value
                if compar_day is None:
                    break

                f_format_day = datetime.datetime.strptime(compar_day, format_str).strftime(format_str)
                dates_positions.append({'valueDate':f_format_day, 'row':now_position})
    
            if len(dates_positions) < 3:
                for i in range(-7, -21, -7):
                    now_position = max_position + i
                    compar_day = spreadsheet.cell(1, now_position).value
                    
                    f_format_day = datetime.datetime.strptime(compar_day, format_str).strftime(format_str)
                    dates_positions.append({'valueDate':f_format_day, 'row':now_position})
                    if len(dates_positions) ==3:
                        break

            return dates_positions
        except:
            return print('except : None')

    def get_old_sheet(self, postionCell, sub_name):
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sub_name)

        start_row = get_column_letter(postionCell-1)
        end_row = get_column_letter(postionCell+4)

        # 範囲指定の形式を修正
        range_string = f'{start_row}1:{end_row}37'
        base_data = spreadsheet.batch_get([range_string])  # batch_getはリスト形式の引数を取る
        return base_data



