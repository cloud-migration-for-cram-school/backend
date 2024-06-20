from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import gspread
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
import datetime
import heapq
import json
from openpyxl.utils import get_column_letter


#fastapiの有効化
app = FastAPI(debug=True)

# FastAPI に CORS ミドルウェアを追加
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # 任意のオリジンからのリクエストを許可
    allow_credentials=True,
    allow_methods=["*"], # 任意の HTTP メソッドを許可
    allow_headers=["*"], # 任意のヘッダーを許可
)

api_path = r"C:\Users\sabax\Documents\backend\apitest-418907-1658bd7509d0.json"
folder_id = '163LOPPPAgKUZXpegrYbBbZkGntFnH5-v'

#google driveに設定
scopes = ['https://www.googleapis.com/auth/drive']

credentials = Credentials.from_service_account_file(
    api_path,
    scopes = scopes
)

class SpreadsheetService:
    def __init__(self, fileID=None):
        credentials = Credentials.from_service_account_file(
            api_path,
            scopes = scopes
        )
        self.credentials = credentials
        self.fileID = fileID
        self.gc = gspread.authorize(credentials)
        #Google Drive APIに接続
        self.drive_service = build('drive', 'v3', credentials=credentials)
        # Google Sheets APIに接続
        self.sheets_service = build('sheets', 'v4', credentials=credentials)
        self.query = f"'{folder_id}' in parents"

    def get_info(self):
        """
        スプレッドシートのファイル名とIDを取得
        """
        try:
            results = self.drive_service.files().list(q=self.query, fields='files(name, id)').execute()
            items = results.get('files', [])
            return items
        except Exception as e:
            print(e)



    def get_worksheet(self):
        """
        sheetidのシート名とIDを辞書として返す
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        worksheets = spreadsheet.worksheets()
        return [{'label': ws.title, 'value': ws.id} for ws in worksheets]


    def get_date_info(self, sheetname):
        """
        日付の入った行の値を取得
        returnは配列
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sheetname)
        compar_day = spreadsheet.row_values(1)
        print(compar_day)


    def get_closest_positions(self, dates_positions, target_date):
        """
        今日から最も日付が近いセルの位置を取得
        dict型 'row': 列番号, 'str_row' : '文字型の列番号(ex ABG etc)'
        """
        format_str = "%m/%d %H:%M"
        differences = [abs(datetime.datetime.strptime(target_date, format_str) - datetime.datetime.strptime(date_position[0], format_str)) for date_position in dates_positions]
        closest_indices = heapq.nsmallest(3, range(len(differences)), key=differences.__getitem__)
        return [(dates_positions[i][1]) for i in closest_indices]



    def closetDataFinder(self, sheetname):
        """
        取得した日付と近い順に３つ値と位置を取得する
        returnはint 位置のみ
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sheetname)

        format_str = "%m/%d %H:%M"
        dates_positions = []
        row = 2

        while True:
            compar_day = spreadsheet.cell(1, row).value
            if compar_day is None:
                break

            f_format_day = datetime.datetime.strptime(compar_day, format_str).strftime(format_str)
            dates_positions.append({'value':f_format_day, 'position':row})
            row += 42  # 4枚スキップ

        # 選択したデータの近辺をさらに取得
        return self.near_compar_date(sheetname=sheetname, row_data=dates_positions)

    def near_compar_date(self, sheetname, row_data):
        """
        日付に関してのデータの取得
        closetDataFinderで得たデータ付近を探索して日付が近いデータの位置を返す
        """
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sheetname)
        dates_positions = []

        format_str = "%m/%d %H:%M"
        now_date = datetime.datetime.now()
        f_now_date = now_date.strftime(format_str)
        try:
            max_position = max(item['position'] for item in row_data)    # dates_positionsの最大値を取得する

            for i in range(0, 35, 7):
                now_position = max_position + i
                compar_day = spreadsheet.cell(1, now_position).value
                if compar_day is None:
                    break

                f_format_day = datetime.datetime.strptime(compar_day, format_str).strftime(format_str)
                dates_positions.append({'valueDate':f_format_day, 'row':now_position})
    
            if len(dates_positions) < 3:
                for i in range(-7, -21, -7):
                    now_position = max_position + i
                    compar_day = spreadsheet.cell(1, now_position).value
                    
                    f_format_day = datetime.datetime.strptime(compar_day, format_str).strftime(format_str)
                    dates_positions.append({'valueDate':f_format_day, 'row':now_position})
                    if len(dates_positions) ==3:
                        break

            return dates_positions
        except:
            return print('except : None')

    def get_old_sheet(self, postionCell, sub_name):
        sheet_url = f'https://docs.google.com/spreadsheets/d/{self.fileID}/edit?usp=sharing'
        spreadsheet = self.gc.open_by_url(sheet_url)
        spreadsheet = spreadsheet.worksheet(sub_name)

        start_row = export_cell_position(value=postionCell-1)
        end_row = export_cell_position(value=postionCell+4)

        # 範囲指定の形式を修正
        range_string = f'{start_row}1:{end_row}37'
        base_data = spreadsheet.batch_get([range_string])  # batch_getはリスト形式の引数を取る
        return base_data




#google drive内のファイル名とIDを取得
@app.get('/search')
def get_search_info():
    sp = SpreadsheetService()
    try:
        items = sp.get_info()
        # Create a list to hold the file names and IDs
        file_info = []
        # Loop through the items and add the file name and ID to the list
        for item in items:
            file_info.append({'label': item['name'], 'value': item['id']})
        return file_info
    except HttpError as error:
        message = f'エラー: {error.content.decode("utf-8")}'
        return message

@app.get('/search/subjects/{sheet_id}')
def get_subjects(sheet_id: str):
    sp = SpreadsheetService(fileID=sheet_id)
    subject = sp.get_worksheet()
    return subject


@app.get('/search/subjects/reports/{subjects_id}}')
def user_info(sheet_id: str, subjects_id: int):
    sp = SpreadsheetService(fileID=sheet_id)
    # 指定されたシートIDに対応するシートの情報を取得
    subjects = sp.get_worksheet()
    # subjects_idを使用して、対応するシートの名前を検索
    for subject in subjects:
        if subject['value'] == subjects_id:
            sheet_name = subject['label']
            break
    else:
        return {"error": "Subject not found"}
    print(f'科目名 : {sheet_name}')
    date_info = sp.closetDataFinder(sheetname=sheet_name)

    # date_infoを大きい順に並び替える
    sorted_date_info = sorted(date_info, key=lambda x: x['row'], reverse=True)
    print(f'sorted date info : {sorted_date_info}')

    positions = [item['row'] for item in sorted_date_info]
    old_sheet = sp.get_old_sheet(postionCell=positions[0], sub_name=sheet_name)
    print(f"old_sheet : {old_sheet}")


def export_cell_position(value):
    # 列のインデックスをA1表記に変換
    column_letter = get_column_letter(value)
    return column_letter




if __name__ == '__main__':
    z = user_info(sheet_id='1CEn2feUeQMfq885PVtyX96-ImOQxeqypeyePHpnPMg4', subjects_id = 1059696948)
    #z = get_subjects(sheet_id='1CEn2feUeQMfq885PVtyX96-ImOQxeqypeyePHpnPMg4')
    print(z)
