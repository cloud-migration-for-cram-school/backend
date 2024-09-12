from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError as APIError
import gspread
import datetime
import heapq
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

# .envファイルから環境変数を読み込む
load_dotenv()

API_PATH = os.getenv('API_PATH')
FOLDER_ID = os.getenv('FOLDER_ID')

#google driveに設定
SCOPES = ['https://www.googleapis.com/auth/drive']

# シートの設定 ... 日付の間隔は7の倍数で入っている
GET_SAMPLE = 1 # 取得するデータの数
SKIP_SHEET_DATA = 42 # スキップするシートの数 42/7 =  4枚スキップする
NEXT_DATA = 35 # 次のデータを取得する数 5枚取得する
PAST_DATA = -21 # 過去のデータを取得する数 3枚取得する
DATETIME_FORMAT = "%m/%d %H:%M"
NOTIFICATION_THRESHOLD = 5  # 空の報告書が5未満になったら通知を送る。

class SpreadsheetService:
    def __init__(self, fileID=None):
        credentials = Credentials.from_service_account_file(
            API_PATH,
            scopes = SCOPES
        )
        self.sheet_url = f'https://docs.google.com/spreadsheets/d/{fileID}/edit?usp=sharing'
        self.gc = gspread.authorize(credentials)
        self.spreadsheet = self.gc.open_by_url(self.sheet_url)

    def get_worksheets(self):
        """
        sheetidのシート名とIDを辞書として返す
        """
        worksheets = self.spreadsheet.worksheets()
        return [{'label': ws.title, 'value': ws.id} for ws in worksheets]

    def get_closest_positions(self, dates_positions, target_date):
        """
        今日から最も日付が近いセルの位置を取得
        dict型 'row': 列番号, 'str_row' : '文字型の列番号(ex ABG etc)'
        """
        differences = [abs(datetime.datetime.strptime(target_date, DATETIME_FORMAT) - datetime.datetime.strptime(date_position[0], DATETIME_FORMAT)) for date_position in dates_positions]
        closest_indices = heapq.nsmallest(GET_SAMPLE, range(len(differences)), key=differences.__getitem__)
        return [(dates_positions[i][1]) for i in closest_indices]

    def find_closest_dates(self, subject_id, start_row):
        """
        線形探索を用いて日付を取得
        引数 subject_id:シートID, start_row:指数探索後の位置
        return valueDate: 日付, row: 位置
        """
        self.sheet = self.spreadsheet.get_worksheet_by_id(subject_id)
        dates_positions = []
        
        while True:
            try:
                compar_day = self.sheet.cell(1, start_row).value
                if compar_day is None:
                    break

                f_format_day = datetime.datetime.strptime(compar_day, DATETIME_FORMAT).strftime(DATETIME_FORMAT)
                dates_positions = [{'value':f_format_day, 'position':start_row}]
                start_row += SKIP_SHEET_DATA
            except APIError as e:
                print(f'find_closest_dates: {e}')
                break
        # 選択したデータの近辺をさらに取得
        return self.find_nearby_dates(dates_positions)

    def find_nearby_dates(self, row_data):
        """
        線形探索
        日付に関してのデータの取得
        find_closest_datesで得たデータから、さらに細かく刻んで最新のデータを取得する
        return valueDate: 日付, row: 位置
        メモ-row_dataはリスト
        """
        empty_cell_count = 0  # 空のセルを数えるカウンター
        dates_positions = 0
        try:
            while empty_cell_count < NOTIFICATION_THRESHOLD:
                position = row_data[0]['position']

                compar_day = self.sheet.cell(1, position).value
                if compar_day:
                    f_format_day = datetime.datetime.strptime(compar_day, DATETIME_FORMAT).strftime(DATETIME_FORMAT)
                    dates_positions = [{'value': f_format_day, 'row': position}]
                else:
                    empty_cell_count += 1

                row_data[0]['position'] += 7  # 1枚スキップ取得

            return dates_positions, ""

        except APIError as e:
            print(f'報告書の残りの枚数が{empty_cell_count}枚です。')
            print(f'範囲外になりました\n{e}')
            return dates_positions, empty_cell_count

        except Exception as e:
            print(f'報告書の残りの枚数が{empty_cell_count}枚です。')
            print(f'except nearcompar date : {e}')
            return dates_positions, empty_cell_count

    def get_old_sheet_data(self, postionCell, subject_id):
        """
        過去のデータを取得する
        """
        self.spreadsheet = self.spreadsheet.get_worksheet_by_id(subject_id)

        start_row = get_column_letter(postionCell-1)
        end_row = get_column_letter(postionCell+4)

        range_string = f'{start_row}1:{end_row}37'
        base_data = self.spreadsheet.batch_get([range_string])  # batch_getはリスト形式の引数を取る
        return base_data

    def find_exponential_dates(self, subject_id, expotent_base, start_row):
        """
        指数探索版のfind_closet_datesを使って日付を取得
        2回指数探索を用いて端までデータを取得し、線形探索(find_nearby_dates)を用いてデータを確定する
        引数はシート名と指数の基数
        return はvalue : 日付, row : 位置 にする予定
        """
        self.sheet = self.spreadsheet.get_worksheet_by_id(subject_id)

        dates_positions = []
        iteration_count = 0  # イテレーションのカウントを追跡

        while True:
            if iteration_count == 0 :
                row = 0
            else:
                row = expotent_base ** iteration_count
            
            try:
                compar_day = self.sheet.cell(1, row + start_row).value
                if compar_day is None and iteration_count==0:
                    # 過去の報告書がない場合, 空の値と位置を返す
                    dates_positions.append({'value':'', 'position':0})
                    break
                elif compar_day is None:
                    break

                f_format_day = datetime.datetime.strptime(compar_day, DATETIME_FORMAT).strftime(DATETIME_FORMAT)
                dates_positions = [{'value':f_format_day, 'position':row+start_row}]
                iteration_count += 1
            except:
                print("except:datafinder-処理の終了")
                break
        return dates_positions
    
    def update_report(self, start_position, data, subject_id):
        try:
            sheet = self.spreadsheet.get_worksheet_by_id(subject_id)
            range_string = f"{get_column_letter(start_position)}1:{get_column_letter(start_position + len(data[0]) - 1)}{len(data)}"

            value_range = {
                "range": range_string,
                "values": data
            }

            sheet.batch_update([value_range])
        except Exception as e:
            print(f"Failed to update report: {e}")
            raise
